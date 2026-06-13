"""
Загрузчик правил исключений из Excel в таблицу exclusion_rules.

Ожидаемый формат Excel (два листа):
  Лист 1 «General» (или первый лист) — исключения для всех застрахованных
  Лист 2 «Family» (или второй лист) — дополнительные исключения для членов семьи

Колонки (можно без заголовка, порядок фиксирован):
  A — Описание исключения (текст из вординга)
  B — Основные коды МКБ-10 (запятая-разделённые коды и диапазоны)
  C — Смежные/дополнительные коды (то же жёсткое исключение, что и B)

CARVEOUT: вытягивается из текста описания регулярными выражениями.
  Грузинские ключевые фразы, влекущие carveout:
    "გარდა ურგენტული" / "ургентн" → 'urgent'
    "გარდა პირველადი" / "первичн" / "диагностик" → 'diagnostic'
    "გარდა პირველი ჯერ" / "первого раза" / "first_test" → 'first_test'

Использование:
    python -m db.loaders.load_exclusions \
        --file db/data/exclusions.xlsx \
        --tenant-id 00000000-0000-0000-0000-000000000001

    # Пересоздать (удалить старые, загрузить заново):
    python -m db.loaders.load_exclusions --file ... --tenant-id ... --replace
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path
from uuid import UUID

import structlog

log = structlog.get_logger()

# Паттерны carveout из текста описания
_CARVEOUT_PATTERNS: list[tuple[str, str]] = [
    # urgent: ургентное / экстренное вмешательство
    (r"გარდა\s+ურგენტ|urgent|ургент|экстренн|emergency", "urgent"),
    # diagnostic: первичная диагностика
    (r"გარდა\s+პირველად|диагностик|diagnostic|პირველადი\s+დიაგ", "diagnostic"),
    # first_test: первое обращение / первый тест
    (r"გარდა\s+პირველი\s+ჯერ|первого\s+раза|first.?test|первое\s+обращен", "first_test"),
]


def _detect_carveout(description: str) -> list[str]:
    """Извлечь список CARVEOUT-условий из текста описания."""
    result: list[str] = []
    desc_lower = (description or "").lower()
    for pattern, label in _CARVEOUT_PATTERNS:
        if re.search(pattern, desc_lower, re.IGNORECASE):
            if label not in result:
                result.append(label)
    return result


def _parse_codes(raw: str) -> list[str]:
    """
    Разобрать строку с кодами/диапазонами МКБ-10.

    Разделители: запятая, точка с запятой, пробел (кроме пробела внутри кода).
    Нормализует en-dash к ASCII дефису.
    """
    if not raw:
        return []
    normalized = (
        str(raw)
        .replace("–", "-")   # en-dash
        .replace("—", "-")   # em-dash
    )
    # Разбиваем по запятой, точке с запятой или нескольким пробелам
    parts = re.split(r"[,;]+|\s{2,}", normalized)
    codes = [p.strip() for p in parts if p.strip()]
    # Отфильтровать явно некорректные (слишком короткие или без буквы)
    return [c for c in codes if len(c) >= 1 and any(ch.isalpha() for ch in c)]


def _load_sheet(ws, scope: str) -> list[dict]:
    """Разобрать лист Excel в список правил."""
    rules = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if not row or not any(row):
            continue
        description = str(row[0] or "").strip()
        if not description:
            continue

        primary_codes = _parse_codes(str(row[1] or "") if len(row) > 1 else "")
        related_codes = _parse_codes(str(row[2] or "") if len(row) > 2 else "")
        # Все коды — одинаково жёсткое исключение (ответ пользователя 2026-06-13)
        all_codes = list(dict.fromkeys(primary_codes + related_codes))  # deduplicate

        carveout = _detect_carveout(description)

        rules.append({
            "scope": scope,
            "description": description,
            "icd10_codes": all_codes,
            "carveout_conditions": carveout,
            "source_row": i,
        })
    return rules


async def _load(file_path: Path, tenant_id: UUID, replace: bool) -> None:
    import openpyxl
    from sqlalchemy import delete
    from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker

    from core.config import get_settings
    from core.models.exclusion import ExclusionRule

    settings = get_settings()
    engine = create_async_engine(settings.database_url, echo=False)
    Session = async_sessionmaker(engine, expire_on_commit=False)

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets = wb.sheetnames

    # Определяем лист для общих и семейных исключений
    # Первый лист → 'all', второй (если есть) → 'family'
    sheet_pairs: list[tuple[object, str]] = []
    for idx, name in enumerate(sheets[:2]):
        scope = "family" if idx == 1 else "all"
        sheet_pairs.append((wb[name], scope))
        log.info("loading_sheet", sheet=name, scope=scope)

    all_rules: list[dict] = []
    for ws, scope in sheet_pairs:
        all_rules.extend(_load_sheet(ws, scope))

    log.info("parsed_rules", total=len(all_rules))

    async with Session() as session:
        if replace:
            await session.execute(
                delete(ExclusionRule).where(ExclusionRule.tenant_id == tenant_id)
            )
            log.info("old_rules_deleted", tenant_id=str(tenant_id))

        from sqlalchemy.dialects.postgresql import insert as pg_insert
        from datetime import timezone, datetime

        for rule in all_rules:
            obj = ExclusionRule(
                tenant_id=tenant_id,
                scope=rule["scope"],
                description=rule["description"],
                icd10_codes=rule["icd10_codes"],
                carveout_conditions=rule["carveout_conditions"],
                source_row=rule["source_row"],
                created_at=datetime.now(tz=timezone.utc),
            )
            session.add(obj)

        await session.commit()
        log.info("exclusion_rules_saved", count=len(all_rules), tenant_id=str(tenant_id))

    await engine.dispose()


def main() -> None:
    parser = argparse.ArgumentParser(description="Load exclusion rules from Excel")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument(
        "--tenant-id",
        default="00000000-0000-0000-0000-000000000001",
        help="Tenant UUID (default: дефолтный тенант)",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Удалить существующие правила тенанта перед загрузкой",
    )
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"ERROR: файл не найден: {file_path}", file=sys.stderr)
        sys.exit(1)

    tenant_id = UUID(args.tenant_id)
    asyncio.run(_load(file_path, tenant_id, args.replace))


if __name__ == "__main__":
    main()
